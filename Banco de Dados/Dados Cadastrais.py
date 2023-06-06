from tkinter import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Dados Cadastrais.xlsx')

ws = wb.active
ws.title = 'Dados Cadastrais'


def save():
    to_append = [enome.get(), ecpf.get(), etelefone.get(), edata.get(), ebairro.get(), ecidade.get(), euf.get()]
    ws.append(to_append)


def exit():
    wb.save('Dados Cadastrais.xlsx')
    janela.destroy()


def search():
    resultado_window = Toplevel(janela)
    resultado_window.title('Resultado da Pesquisa')

    # Ler o arquivo Excel usando pandas
    df = pd.read_excel('Dados Cadastrais.xlsx')

    # Criar uma tabela usando uma caixa de texto
    resultado_text = Text(resultado_window, width=100, height=20)
    resultado_text.pack()

    # Adicionar uma barra de rolagem à caixa de texto
    scrollbar = Scrollbar(resultado_window)
    scrollbar.pack(side=RIGHT, fill=Y)
    resultado_text.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=resultado_text.yview)

    # Formatar os dados para exibição
    colunas = df.columns.tolist()
    linhas = df.values.tolist()

    # Inserir os cabeçalhos das colunas na tabela
    resultado_text.insert(END, "\t".join(colunas) + "\n")

    # Inserir os dados das linhas na tabela
    for linha in linhas:
        resultado_text.insert(END, "\t".join(str(valor) for valor in linha) + "\n")

    # Impedir que o usuário edite a tabela
    resultado_text.configure(state='disabled')



janela = Tk()
janela.title('Cadastro')
janela.config(borderwidth=10)
janela.option_add('*Font', 'Arial 8 bold')

dados = LabelFrame(janela, text='Dados Pessoais', font='Arial 14 italic')
dados.grid(row=0, column=0, sticky='we', columnspan=5)
dados.config(padx=10, pady=10)

txnome = Label(dados, text='Nome:', anchor='e')
txnome.grid(row=0, column=0, sticky='we')
txcpf = Label(dados, text='CPF:', anchor='e')
txcpf.grid(row=1, column=0, sticky='we')
txtelefone = Label(dados, text='Telefone:', anchor='e')
txtelefone.grid(row=1, column=2, sticky='we')
txdata = Label(dados, text='Data de Nasc:')
txdata.grid(row=1, column=4, sticky='we')

enome = Entry(dados, width=100)
enome.grid(row=0, column=1, columnspan=5)
ecpf = Entry(dados, width=25)
ecpf.grid(row=1, column=1)
etelefone = Entry(dados, width=25)
etelefone.grid(row=1, column=3)
edata = Entry(dados, width=25)
edata.grid(row=1, column=5)

endereco = LabelFrame(janela, text='Endereço', font='Arial 14 italic')
endereco.grid(row=1, column=0, sticky='we', columnspan=5)
endereco.config(pady=10, padx=10)

txrua = Label(endereco, text='Rua:', anchor='e')
txrua.grid(row=0, column=0, sticky='we')
txn = Label(endereco, text='Nº:', anchor='e')
txn.grid(row=0, column=4, sticky='we')
txbairro = Label(endereco, text='Bairro:', anchor='e')
txbairro.grid(row=1, column=0, sticky='we')
txcidade = Label(endereco, text='Cidade:', anchor='e')
txcidade.grid(row=1, column=2, sticky='we')
txuf = Label(endereco, text='UF:', anchor='e')
txuf.grid(row=1, column=4, sticky='we')

erua = Entry(endereco, width=80)
erua.grid(row=0, column=1, columnspan=3)
en = Entry(endereco, width=15)
en.grid(row=0, column=5)
ebairro = Entry(endereco, width=35)
ebairro.grid(row=1, column=1)
ecidade = Entry(endereco, width=35)
ecidade.grid(row=1, column=3)
euf = Entry(endereco, width=15)
euf.grid(row=1, column=5)

botoes = LabelFrame(janela)
botoes.grid(row=2, column=0, sticky='we')
botoes.config(borderwidth=0)

gravar = Button(botoes, text='Gravar Dados', command=save)
gravar.grid(row=2, column=0)
sair = Button(botoes, text='Sair', command=exit)
sair.grid(row=2, column=1)
pesq = Button(botoes, text='Pesquisar', command=search)
pesq.grid(row=2, column=2)

janela.mainloop()
